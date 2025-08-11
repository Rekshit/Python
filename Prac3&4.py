'''
import numpy as num
import statistics as std
data1 = [13,16,13,19,15]
data2 = [23,27,23,28,25]

mean_1 = num.mean(data1)
mean_2 = num.mean(data2)

variance_1 = num.var(data1)
variance_2 = num.var(data2)
print("The variance of data 1 is: ", variance_1)
print("The variance of data 2 is: ", variance_2)

n1 = len(data1)
n2 = len(data2)

combined_mean = ((n1*mean_1)+(n2*mean_2))/(n1+n2)
print("\nThe combined mean of both data is: ", combined_mean)

d1 = (mean_1 - combined_mean)*(mean_1 - combined_mean)
d2 = (mean_2 - combined_mean)*(mean_2 - combined_mean)

comb_var =((n1*d1*variance_1)+(n1*d2*variance_2))/(n1+n2)
print("\nThe combined variance of both data is: ", comb_var)

sd_1 = std.stdev(data1)
sd_2 = std.stdev(data2)

coeff_var1 = (sd_1/mean_1)*100
coeff_var2 = (sd_2/mean_2)*100
print("\nThe coefficient of variance of data 1 is: ", coeff_var1)
print("The coefficient of variance of data 2 is: ", coeff_var2)
'''
'''
import seaborn as sns
import scipy
from scipy.stats import norm
import matplotlib.pyplot as plt

# Parameters
mean = 50
sd = 1

# Create a normal distribution
n_data = norm(loc=mean, scale=sd)

# Generate a random sample of size 10
sample = n_data.rvs(size=10)

# Compute PDF and CDF values for the sample
pdf_values = n_data.pdf(sample)
cdf_values = n_data.cdf(sample)

# Plotting the sample distribution with KDE
sns.histplot(sample, kde=True, color="green")  # distplot is deprecated
plt.title("Histogram with KDE")

# Show the plot
plt.show()
'''
'''
import seaborn as sns
import scipy
from scipy.stats import norm
import matplotlib.pyplot as plt

# Parameters
mean = 50
sd = 1

# Create a normal distribution
n_data = norm(loc=mean, scale=sd)

# Generate a random sample of size 10
sample = n_data.rvs(size=10)

# Compute PDF and CDF values for the sample
pdf_values = n_data.pdf(sample)
cdf_values = n_data.cdf(sample)

# Plotting the sample distribution with KDE
sns.histplot(sample, kde=True, color="green")

# Show the plot
plt.show()
'''
'''
import seaborn as sns
import scipy
from scipy.stats import norm
import matplotlib.pyplot as plt

# Parameters
mean = 50
sd = 1

# Create a normal distribution
n_data = norm(loc=mean, scale=sd)

# Generate a random sample of size 5
sample = n_data.rvs(size=5)

# Compute PDF and CDF values for the sample
pdf_values = n_data.pdf(sample)
cdf_values = n_data.cdf(sample)

print("\nThe pdf values of generated data is: ", pdf_values )
print("\nThe cdf values of generated data is: ", cdf_values )
print("\n")
# Plotting the sample distribution with KDE
sns.histplot(sample, kde=True, color="green")

# Show the plot
plt.show()
'''
'''
import seaborn as sns
import scipy
from scipy.stats import norm
import matplotlib.pyplot as plt
from scipy.stats import binom

# n=no of trials
# p= probability of success
# pmf=Probability Mass Function
b_dist = binom (n=10, p=0.5)
pfs = b_dist.pmf(5)
print("\n",pfs)
'''
import openpyxl
import csv
from openpyxl import Workbook, load_workbook

# Write data to Excel file
def write_to_excel(data, filename_xls):
    wb = Workbook()
    sheet = wb.active
    for row in data:
        sheet.append(row)
    wb.save(filename_xls)
    print("Data written to Excel successfully")

# Write data to CSV file
def write_to_csv(data, filename_csv):
    with open(filename_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)
    print("Data written to CSV successfully")

import csv
from openpyxl import load_workbook

# Read from Excel file
def read_from_excel(fileb.xlsx):
    data = []
    wb = load_workbook(fileb.xlsx)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

# Read from CSV file
def read_from_csv(filea.csv):
    data = []
    with open(filea.csv, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            data.append(row)
    return data

# --- Main Execution ---

# Change these to your actual file names
excel_filename = "your_excel_file.xlsx"
csv_filename = "your_csv_file.csv"

# Read and display Excel data
excel_data = read_from_excel(excel_filename)
print("\nData from Excel file:")
for row in excel_data:
    print(row)

# Read and display CSV data
csv_data = read_from_csv(csv_filename)
print("\nData from CSV file:")
for row in csv_data:
    print(row)