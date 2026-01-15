import openpyxl
import csv
from openpyxl import Workbook, load_workbook

# Sample data
data = [
    ["Roll No", "Name", "Marks"],
    [1, "Rekshit", 85],
    [2, "Robin", 90],
    [3, "sahil", 78]
]

# Write to Excel file
def write_to_excel(data, filename_xls):
    wb = Workbook()
    sheet = wb.active
    for row in data:
        sheet.append(row)
    wb.save(filename_xls)
    print(f"Excel file '{filename_xls}' created successfully.")

# Write to CSV file
def write_to_csv(data, filename_csv):
    with open(filename_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)
    print(f"CSV file '{filename_csv}' created successfully.")

# Read from Excel file
def read_from_excel(filename_xls):
    wb = load_workbook(filename_xls)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

# Read from CSV file
def read_from_csv(filename_csv):
    data = []
    with open(filename_csv, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            data.append(row)
    return data

# --- Main Execution ---
excel_filename = "r.xlsx"
csv_filename = "reki.csv"

# Create files
write_to_excel(data, excel_filename)
write_to_csv(data, csv_filename)

# Read and display data
excel_data = read_from_excel(excel_filename)
csv_data = read_from_csv(csv_filename)

print("\nData from Excel file:")
for row in excel_data:
    print(row)

print("\nData from CSV file:")
for row in csv_data:
    print(row)